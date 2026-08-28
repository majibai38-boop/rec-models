# Copyright 2025. Huawei Technologies Co.,Ltd. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#    http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ==============================================================================

"""Tensor comparison helpers used by optional precision checks."""

import torch
from typing import Any, Dict, List, Tuple
import os
import collections

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from .logger import logger


def calculate_errors(
        tensor1: torch.Tensor, tensor2: torch.Tensor
) -> Dict[str, float]:
    """
    计算两个扁平化张量之间的平均相对误差和平均绝对误差。
    (可稳健处理 inf 和 nan)
    返回一个包含两种误差的字典。
    """
    tensor1_flat = tensor1.flatten().to(device="cpu", dtype=torch.float64)
    tensor2_flat = tensor2.flatten().to(device="cpu", dtype=torch.float64)

    if tensor1_flat.shape != tensor2_flat.shape:
        return {"relative": -1.0, "absolute": -1.0}  # 返回-1.0表示形状不匹配

    # 1. 找到完全相等的位置（包括 inf == inf），这些位置的误差为 0
    equal_mask = (tensor1_flat == tensor2_flat) | \
                 (torch.isnan(tensor1_flat) & torch.isnan(tensor2_flat))

    # 初始化误差张量
    relative_error = torch.zeros_like(tensor1_flat, dtype=torch.float64)
    absolute_error = torch.zeros_like(tensor1_flat, dtype=torch.float64)

    # 2. 找到不相等的位置
    unequal_mask = ~equal_mask

    # 如果没有不相等的值，误差为0
    if not torch.any(unequal_mask):
        return {"relative": 0.0, "absolute": 0.0}

    t1_unequal = tensor1_flat[unequal_mask]
    t2_unequal = tensor2_flat[unequal_mask]

    # 3. 在不相等的值中，处理涉及 inf 或 nan 的情况
    inf_nan_mask = torch.isinf(t1_unequal) | torch.isinf(t2_unequal) | \
                   torch.isnan(t1_unequal) | torch.isnan(t2_unequal)

    # 初始化不相等部分的误差张量
    unequal_rel_errors = torch.zeros_like(t1_unequal, dtype=torch.float64)
    unequal_abs_errors = torch.zeros_like(t1_unequal, dtype=torch.float64)

    # 将inf/nan位置的误差设置为inf
    unequal_rel_errors[inf_nan_mask] = torch.inf
    unequal_abs_errors[inf_nan_mask] = torch.inf

    # 4. 对剩下的有限且不相等的值计算误差
    finite_mask = ~inf_nan_mask
    if torch.any(finite_mask):
        t1_finite = t1_unequal[finite_mask]
        t2_finite = t2_unequal[finite_mask]

        # 新增：计算绝对误差
        finite_abs_errors = torch.abs(t1_finite - t2_finite)
        unequal_abs_errors[finite_mask] = finite_abs_errors

        # 计算相对误差
        epsilon = torch.finfo(torch.float64).eps
        denominator = torch.abs(t1_finite)
        denominator = torch.where(denominator < epsilon, epsilon, denominator)
        finite_rel_errors = torch.abs(t1_finite - t2_finite) / denominator
        unequal_rel_errors[finite_mask] = finite_rel_errors

    # 将计算出的误差放回主误差张量
    relative_error[unequal_mask] = unequal_rel_errors
    absolute_error[unequal_mask] = unequal_abs_errors

    # 5. 计算最终的平均误差。如果存在 inf，结果就是 inf。
    return {
        "relative": relative_error.mean().item(),
        "absolute": absolute_error.mean().item()
    }


def extract_data_by_type(item: Any) -> Tuple[List[torch.Tensor], List[str]]:
    """
    递归地从项目中提取所有数值和非数值数据。
    返回一个元组：(数值张量列表, 非数值字符串列表)。
    """
    tensors = []
    non_numerics = []

    if isinstance(item, torch.Tensor):
        tensors.append(item)
    elif isinstance(item, (int, float)):
        tensors.append(torch.tensor([item], dtype=torch.float64))
    elif isinstance(item, (list, tuple)):
        for sub_item in item:
            sub_tensors, sub_non_numerics = extract_data_by_type(sub_item)
            tensors.extend(sub_tensors)
            non_numerics.extend(sub_non_numerics)
    elif isinstance(item, dict):
        for _, value in sorted(item.items()):  # 排序以保证确定性
            sub_tensors, sub_non_numerics = extract_data_by_type(value)
            tensors.extend(sub_tensors)
            non_numerics.extend(sub_non_numerics)
    elif item is not None:
        # 其他所有类型都视为非数值，并转换为字符串
        non_numerics.append(str(item))

    return tensors, non_numerics


def compare_items(
        item1: Any,
        item2: Any,
        rel_error_threshold: float = 2e-3,
        abs_error_threshold: float = 2e-3,
) -> Dict[str, Any]:
    """
    比较两个项目，分别处理数值和非数值部分，
    并将详细比较结果返回到字典中。
    """
    result = {}

    # 1. 类型和None值检查 (最高优先级)
    if type(item1) is not type(item2):
        result["Comparison"] = (
            f"Type Mismatch: {type(item1).__name__} vs {type(item2).__name__}"
        )
        return result
    if item1 is None and item2 is None:
        result["Comparison"] = "Both are None"
        return result
    if item1 is None or item2 is None:
        result["Comparison"] = "One is None"
        return result

    # 2. 提取数值和非数值数据
    numerics1, non_numerics1 = extract_data_by_type(item1)
    numerics2, non_numerics2 = extract_data_by_type(item2)

    comparison_parts = []

    # 3. 比较数值部分
    if numerics1 or numerics2:
        if len(numerics1) != len(numerics2):
            comparison_parts.append("Numeric: Structure Mismatch (count differs)")
            result["Avg_Rel_Error"] = "N/A (count mismatch)"
            result["Avg_Abs_Error"] = "N/A (count mismatch)"  # 新增
        else:
            # 检查形状
            shape_mismatches = [
                f"Idx {i}: {t1.shape} vs {t2.shape}"
                for i, (t1, t2) in enumerate(zip(numerics1, numerics2))
                if t1.shape != t2.shape
            ]
            if shape_mismatches:
                comparison_parts.append("Numeric: Shape Mismatch")
                result["Avg_Rel_Error"] = "Shape Mismatch"
                result["Avg_Abs_Error"] = "Shape Mismatch"
                result["Shape_Details"] = "; ".join(shape_mismatches)
            else:
                # 计算聚合误差
                try:
                    all_tensors1 = torch.cat([t.flatten() for t in numerics1])
                    all_tensors2 = torch.cat([t.flatten() for t in numerics2])
                    errors = calculate_errors(all_tensors1, all_tensors2)

                    total_rel_error = errors["relative"]
                    total_abs_error = errors["absolute"]

                    result["Avg_Rel_Error"] = total_rel_error
                    result["Avg_Abs_Error"] = total_abs_error

                    # 同时检查相对和绝对误差
                    rel_exceeded = total_rel_error > rel_error_threshold
                    abs_exceeded = total_abs_error > abs_error_threshold

                    if rel_exceeded or abs_exceeded:
                        error_msgs = []
                        if rel_exceeded:
                            error_msgs.append(f"Rel Err > {rel_error_threshold:.0e}")
                        if abs_exceeded:
                            error_msgs.append(f"Abs Err > {abs_error_threshold:.0e}")
                        comparison_parts.append(f"Numeric: Not Equal ({'; '.join(error_msgs)})")
                    else:
                        comparison_parts.append("Numeric: Equal")
                except Exception as e:
                    error_msg = f"Error during cat/calc: {e}"
                    result["Avg_Rel_Error"] = error_msg
                    result["Avg_Abs_Error"] = error_msg  # 新增
                    comparison_parts.append("Numeric: Error during calculation")

    # 4. 比较非数值部分
    if non_numerics1 or non_numerics2:
        if non_numerics1 == non_numerics2:
            comparison_parts.append("Non-Numeric: Equal")
        else:
            comparison_parts.append("Non-Numeric: Not Equal")
            result["NonNumeric_Details"] = (
                f"'{' | '.join(non_numerics1)}' vs '{' | '.join(non_numerics2)}'"
            )

    # 5. 合并最终的比较摘要
    if not comparison_parts:
        result["Comparison"] = "No data to compare"
    else:
        result["Comparison"] = "; ".join(comparison_parts)

    return result

def style_excel_sheet(
        writer: pd.ExcelWriter,
        df: pd.DataFrame,
        rel_error_threshold: float = 1e-5,
        abs_error_threshold: float = 1e-5,  # 新增
):
    """ 应用样式到生成的Excel工作表，使用 DifferentialStyle 以确保高亮生效。"""
    if df.empty or len(df) < 1:
        return

    workbook = writer.book
    worksheet = writer.sheets["Comparison"]

    # --- 1. 基于节点类型为行上色 ---
    if "Node Type" in df.columns:
        color_map = {
            "call_module": "AEC6CF",  # 灰蓝色
            "call_function": "B9F2B5",  # 浅绿色
            "get_attr": "DCD0FF",  # 淡紫色
            "placeholder": "FFF6A6",  # 浅黄色
            "output": "E5E4E2",  # 浅灰色
        }
        backup_palette = [
            "C3B1E1",
            "ACE5EE",
            "E0FFFF",
            "F0F8FF",
            "98FB98",
            "B0C4DE"
        ]
        fills = {
            ntype: PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )
            for ntype, hex_color in color_map.items()
        }
        next_backup_idx = 0
        for idx, node_type in enumerate(df["Node Type"].dropna()):
            if node_type not in fills:
                color = backup_palette[next_backup_idx % len(backup_palette)]
                fills[node_type] = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                next_backup_idx += 1
            fill_to_apply = fills.get(node_type)
            if fill_to_apply:
                for cell in worksheet[idx + 2]:
                    cell.fill = fill_to_apply

    # --- 2. 错误高亮 ---
    red_font = Font(color="9C0006")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    dxf = DifferentialStyle(font=red_font, fill=red_fill)

    error_cols = [
        col
        for col in df.columns
        if "Error" in col or "Mismatch" in col or "Not Equal" in col
    ]
    keywords_to_check = ["Mismatch", "Error", "N/A", "Not Equal"]

    for col_name in error_cols:
        if col_name not in df.columns:
            continue

        threshold = abs_error_threshold if "Abs_Error" in col_name else rel_error_threshold

        col_idx = df.columns.get_loc(col_name) + 1
        col_letter = get_column_letter(col_idx)
        range_to_format = f"{col_letter}2:{col_letter}{worksheet.max_row}"
        first_cell = f"{col_letter}2"

        text_checks = [
            f'ISNUMBER(SEARCH("{key}", {first_cell}))'
            for key in keywords_to_check
        ]
        formula = f'=OR(AND(ISNUMBER({first_cell}), {first_cell}>{threshold}), {", ".join(text_checks)})'

        rule = Rule(type="expression", formula=[formula], stopIfTrue=True, dxf=dxf)
        worksheet.conditional_formatting.add(range_to_format, rule)

    # --- 3. 添加边框 ---
    thin_border_side = Side(style="thin", color="000000")
    thin_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )
    for row in worksheet.iter_rows(
            min_row=1, max_row=worksheet.max_row, max_col=worksheet.max_column
    ):
        for cell in row:
            cell.border = thin_border

    # --- 4. 自动调整列宽和格式 ---
    for i, column_cells in enumerate(worksheet.columns):
        is_error_col = "Error" in df.columns[i]
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            if is_error_col and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00E+00"
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                logger.info(e)
        adjusted_width = min(max(max_length + 2, 12), 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    worksheet.freeze_panes = "A2"


def compare_captured_data(
        data1,
        data2,
        output_excel_path: str,
        rel_error_threshold: float = 2e-3,
        abs_error_threshold: float = 2e-3,
        print_result=False
):
    """
    加载两个捕获的数据文件，
    并将详细的比较结果保存到格式化的Excel文件中。
    """
    if not isinstance(data1, dict) or not isinstance(data2, dict):
        logger.info("错误：一个或两个输入文件不包含预期的字典数据。")
        return

    # 2. 准备节点键列表 (逻辑不变)
    keys1 = set(data1.keys())
    keys2 = set(data2.keys())
    node_keys_ordered = (
        list(data1.keys())
        if isinstance(data1, collections.OrderedDict)
        else sorted(list(keys1))
    )
    nodes_only_in_file2 = sorted(list(keys2 - keys1))
    all_node_keys = node_keys_ordered + nodes_only_in_file2

    comparison_results = []

    # 3. 遍历节点进行比较
    for node_name in all_node_keys:
        logger.info(f"正在比较节点: {node_name}")
        node_data1 = data1.get(node_name)
        node_data2 = data2.get(node_name)
        result_row = {"Node Name": node_name}

        if node_name not in keys1:
            result_row["Comparison_Status"] = (
                f"Node only in gpu data"
            )
        elif node_name not in keys2:
            result_row["Comparison_Status"] = (
                f"Node only in npu data"
            )
        else:
            result_row["Comparison_Status"] = "Present in Both"

        if node_data1 and "type" in node_data1:
            result_row["Node Type"] = node_data1.get("type")
        elif node_data2 and "type" in node_data2:
            result_row["Node Type"] = node_data2.get("type")

        fields_to_compare = set()
        if node_data1:
            fields_to_compare.update(node_data1.keys())
        if node_data2:
            fields_to_compare.update(node_data2.keys())

        for field in sorted(list(fields_to_compare)):
            if field in ["type"]:
                continue

            item1 = node_data1.get(field) if node_data1 else None
            item2 = node_data2.get(field) if node_data2 else None

            comparison = compare_items(
                item1,
                item2,
                rel_error_threshold=rel_error_threshold,
                abs_error_threshold=abs_error_threshold,
            )

            for key, value in comparison.items():
                result_row[f"{field.capitalize()}_{key}"] = value

        comparison_results.append(result_row)

    # 4. 保存到Excel
    if not comparison_results:
        logger.info("没有可比较的数据。")
        return
    if print_result:
        logger.info(f"\ncomparison_results: \n{comparison_results}")
    df = pd.DataFrame(comparison_results)

    cols_order = ["Node Name", "Node Type", "Comparison_Status"]
    other_cols = sorted([col for col in df.columns if col not in cols_order])
    final_cols = [col for col in cols_order if col in df.columns] + other_cols
    df = df[final_cols]

    try:
        with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Comparison")
            style_excel_sheet(
                writer,
                df,
                rel_error_threshold=rel_error_threshold,
                abs_error_threshold=abs_error_threshold,
            )
        logger.info(f"比较结果已成功保存到 '{output_excel_path}'")
    except Exception as e:
        logger.info(f"保存Excel文件时发生错误：{e}")


# --- 主程序入口 ---
if __name__ == "__main__":
    model_name = "hllm"

    gpu_infer_output_file = ""
    npu_infer_output_file = ""
    excel_output_file = f"./{model_name}_l20_a5.xlsx"
    gpu_data = torch.load(gpu_infer_output_file, map_location="cpu", weights_only=False)
    npu_data = torch.load(npu_infer_output_file, map_location="cpu", weights_only=False)
    logger.info(
        f"成功从 '{os.path.basename(gpu_infer_output_file)}' 和 '{os.path.basename(npu_infer_output_file)}' 加载数据。"
    )
    compare_captured_data(
        gpu_data,
        npu_data,
        excel_output_file,
        rel_error_threshold=2e-3,  # 相对误差阈值
        abs_error_threshold=2e-3,  # 绝对误差阈值 (可根据需要调整)
        print_result=True
    )
